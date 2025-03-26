import datetime
import os
from openpyxl import Workbook, load_workbook

class AttendanceTracker:
    def __init__(self):
        self.uid_to_name = {
            "0285212673": "Максим Бармен",
            "0285212674": "Максим Официант",
            "0285212675": "Катя",
            "0285212677": "Серафима",
            "0285212707": "Ева",
            "0285212708": "София",
            "0285212709": "Алина",
            "0285212710": "Люба",
            "0285212711": "Вовичк"
        }
        self.excel_file = "attendance.xlsx"
        if not os.path.exists(self.excel_file):
            self.create_excel_file()

    def create_excel_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Дата", "UID", "Имя", "Статус", "Время"])
        wb.save(self.excel_file)

    def log_attendance(self, uid):
        current_time = datetime.datetime.now()
        name = self.uid_to_name.get(uid, "Неизвестный")
        status = "пришел" if self.is_odd_entry(uid) else "ушел"
        log_message = f"{name} (UID: {uid}) {status} в {current_time}"
        self.write_to_excel(uid, name, current_time, status)
        return log_message

    def is_odd_entry(self, uid):
        wb = load_workbook(self.excel_file)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == uid:
                count += 1
        return count % 2 == 0

    def write_to_excel(self, uid, name, current_time, status):
        wb = load_workbook(self.excel_file)
        ws = wb.active
        ws.append([current_time.strftime("%Y-%m-%d"), uid, name, status, current_time.strftime("%H:%M:%S")])
        wb.save(self.excel_file)

    def get_attendance(self, date):
        date_str = date.strftime("%Y-%m-%d")
        wb = load_workbook(self.excel_file)
        ws = wb.active
        attendance_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == date_str:
                attendance_data.append(f"{row[2]} (UID: {row[1]}) {row[3]} в {row[4]}\n")
        return attendance_data

    def calculate_hours(self, month):
        wb = load_workbook(self.excel_file)
        ws = wb.active
        attendance_log = {uid: [] for uid in self.uid_to_name.keys()}
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, uid, name, status, time = row
            if date.startswith(month):
                timestamp = datetime.datetime.strptime(f"{date} {time}", "%Y-%m-%d %H:%M:%S")
                attendance_log[uid].append(timestamp)
        hours_worked = {}
        for uid, times in attendance_log.items():
            total_seconds = 0
            for i in range(0, len(times), 2):
                if i + 1 < len(times):
                    total_seconds += (times[i + 1] - times[i]).total_seconds()
            hours_worked[uid] = total_seconds / 3600  # Convert seconds to hours
        return hours_worked

    def edit_attendance(self, date, old_entry, new_entry):
        date_str = date.strftime("%Y-%m-%d")
        wb = load_workbook(self.excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == date_str and row[1].value == old_entry.split(" (UID: ")[1].split(")")[0]:
                row[2].value = new_entry.split(" (UID: ")[0]
                row[3].value = new_entry.split(" (UID: ")[1].split(")")[0]
                row[4].value = new_entry.split(" в ")[1]
        wb.save(self.excel_file)